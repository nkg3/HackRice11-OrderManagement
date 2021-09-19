import firebase_admin
from firebase_admin.credentials import Certificate
from firebase_admin import db
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path

DEBUG: bool = False
PRINT_DATA: bool = False


"""
Call init_database first before anything!
"""


def init_database(url: str, cred_path: str):
    cred = Certificate(cred_path)
    default_app = firebase_admin.initialize_app(cred, {'databaseURL': url})


def set_with_dict(data: dict, dir_ref: str = "/"):
    reference = db.reference(dir_ref)
    reference.set(data)
    if DEBUG:
        print("[Database set] source is dict, reference dir: " + dir_ref)
    if PRINT_DATA:
        print(data)


def update(dir_ref: str, data: dict):
    reference = db.reference(dir_ref)
    reference.update(data)
    if DEBUG:
        print("[Database update] reference dir: " + dir_ref)
    if PRINT_DATA:
        print(data)


def remove(dir_ref: str):
    reference = db.reference(dir_ref)
    reference.delete()


def set_with_excel(path: str, sheet_name: str, row_range: tuple, col_range: tuple, dir_ref: str = "/"):
    reference = db.reference(dir_ref)
    parser: ExcelParser = ExcelParser(Path(path))
    data = parser.parse_sheet(sheet_name, row_range, col_range)
    reference.set(data)
    if DEBUG:
        print("[Database set] sheet name: " + sheet_name + ", reference dir: " + dir_ref)
    if PRINT_DATA:
        print(data)


def get_under_directory(dir_ref: str, child_key: str = "", child_value: str = "") -> object:
    """
    get values from database
    Functionalities:
        get a dictionary of all children under the directory (THIS DEFINITELY WORKS RIGHT NOW)
        get one child given the child_key and child_value (specify child_key and child_value)
        get an OrderedDict of children given a key to index with (specify indexing_key)
    :param dir_ref: directory of the data
    :param child_key: key of child to index with
    :param child_value: value of child
    :return: A dict if only given dir_ref. An OrderedDict if given other parameters
    """
    reference = db.reference(dir_ref)
    if not child_value == "":
        return reference.order_by_child(child_key).equal_to(child_value).get()
    if (not child_key == "") and child_value == "":
        return reference.order_by_child(child_key).get()
    return reference.get()


def remove_symbol(word: str, symbol: str):
    if word[-1] == symbol:
        return word[:-1]
    return word


def demo_parse_excel_to_db(path: str):
    database_url = 'https://hackrice11-ordermanageme-327b0-default-rtdb.firebaseio.com/'
    cred_path = Path("../DataFiles/database_private_key.json")  # path to database credential
    excel_path = Path(path)  # path to excel sheets

    init_database(database_url, cred_path)
    set_with_excel(excel_path, "Equipment Details", (2, 12), (2, 9), "/Equipments")
    set_with_excel(excel_path, "Worker Details", (1, 11), (2, 7), "/Workers")
    set_with_excel(excel_path, "Facility Details", (2, 7), (2, 6), "/Facilities")
    set_with_excel(excel_path, "Work Order Examples", (2, 32), (2, 12), "/WorkOrders")


class ExcelParser:
    __wb: Workbook
    __sheet_names: list

    def __init__(self, path: Path):
        self.__wb = openpyxl.load_workbook(path)
        self.__sheet_names = self.__wb.sheetnames

    def parse_sheet(self, sheet_name: str, row_range: tuple, col_rage: tuple) -> dict:
        if sheet_name not in self.__sheet_names:
            raise KeyError("sheet name doesn't exist")

        data: dict = {}
        sheet: Worksheet = self.__wb[sheet_name]
        rows = sheet.iter_rows(row_range[0], row_range[1], col_rage[0], col_rage[1])
        keys: list = []

        for i, row in enumerate(rows):
            if i == 0:
                for cell in row:
                    keys.append(remove_symbol(str(cell.value), "#"))
            else:
                identifier: str = remove_symbol(str(row[0].value), "#")
                data[identifier] = {}
                for j in range(len(row)):
                    data[identifier][keys[j]] = str(row[j].value)

        return data