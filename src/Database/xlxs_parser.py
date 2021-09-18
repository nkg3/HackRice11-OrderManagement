import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path


def remove_symbol(word: str, symbol: str):
    if word[-1] == symbol:
        return word[:-1]
    return word


class ExcelParser:
    __wb: Workbook
    __sheet_names: list

    def __init__(self, path: Path):
        self.__wb = openpyxl.load_workbook(path)
        self.__sheet_names = self.__wb.sheetnames

    def parse_sheet(self, sheet_name: str, row_range: tuple, col_rage: tuple) -> dict:
        if sheet_name not in self.__sheet_names:
            raise KeyError("sheet name doesn't exist")

        data: dict = {}
        sheet: Worksheet = self.__wb[sheet_name]
        rows = sheet.iter_rows(row_range[0], row_range[1], col_rage[0], col_rage[1])
        keys: list = []

        for i, row in enumerate(rows):
            if i == 0:
                for cell in row:
                    keys.append(remove_symbol(str(cell.value), "#"))
            else:
                identifier: str = remove_symbol(str(row[0].value), "#")
                data[identifier] = {}
                for j in range(len(row)):
                    data[identifier][keys[j]] = str(row[j].value)

        return data
